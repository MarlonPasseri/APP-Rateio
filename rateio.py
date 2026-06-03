"""
Núcleo de cálculo de rateio por GP (Centro de Custo).

Atualmente implementa o rateio de PLANO DE SAÚDE.
A estrutura foi pensada para receber, no futuro, os rateios de
FÉRIAS, RESCISÃO e 13º reaproveitando a leitura da planilha TS.

Planilha TS (base): aba com as colunas
    Id Colaborador | Nome Colaborador | Mês | GP | Horas Trabalhadas | Proporção de Hora
(a coluna "Horas Mês"/"Id" pode existir e é ignorada para o cálculo).
"""

from __future__ import annotations

import re
import unicodedata
from datetime import datetime, date

import openpyxl


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def _norm(texto) -> str:
    """Normaliza texto de cabeçalho: minúsculo, sem acento, sem espaços extras."""
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )
    s = re.sub(r"\s+", " ", s)
    return s


# Sinônimos aceitos para cada coluna lógica da TS.
_COLUNAS = {
    "id": ["id colaborador", "id do colaborador", "id colab", "matricula"],
    "nome": ["nome colaborador", "nome do colaborador", "colaborador", "nome"],
    "mes": ["mes", "mes referencia", "competencia", "mes/ano"],
    "gp": ["gp", "centro de custo", "cc", "numero gp"],
    "horas": ["horas trabalhadas", "horas trab", "horas"],
    "proporcao": [
        "proporcao de hora",
        "proporcao da hora",
        "proporcao das horas",
        "proporcao",
    ],
}

# Colunas obrigatórias para que uma aba seja reconhecida como a planilha TS.
_OBRIGATORIAS = ["nome", "mes", "gp", "horas", "proporcao"]


def _mes_key(valor) -> str | None:
    """Converte um valor de célula de mês em chave 'AAAA-MM'."""
    if valor is None:
        return None
    if isinstance(valor, (datetime, date)):
        return f"{valor.year:04d}-{valor.month:02d}"
    s = str(valor).strip()
    # Tenta formatos comuns: 2025-01, 01/2025, 2025-01-01, jan/25 não suportado.
    m = re.match(r"^(\d{4})[-/](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.match(r"^(\d{1,2})[-/](\d{4})", s)
    if m:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"
    return None


def _to_float(valor) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace("R$", "").replace(" ", "")
    # Formato brasileiro: 1.234,56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# Leitura da planilha TS
# ---------------------------------------------------------------------------
class PlanilhaTS:
    """Representa a planilha TS já carregada em memória."""

    def __init__(self, linhas: list[dict], aba: str):
        self.linhas = linhas  # cada item: id, nome, mes_key, gp, horas, proporcao
        self.aba = aba

    @property
    def meses(self) -> list[str]:
        return sorted({l["mes_key"] for l in self.linhas})

    def colaboradores(self, mes_key: str) -> list[dict]:
        """Colaboradores distintos que possuem horas no mês informado."""
        vistos = {}
        for l in self.linhas:
            if l["mes_key"] != mes_key:
                continue
            chave = l["id"] or l["nome"]
            if chave not in vistos:
                vistos[chave] = {"id": l["id"], "nome": l["nome"]}
        return sorted(vistos.values(), key=lambda x: (x["nome"] or ""))


def _achar_cabecalho(ws):
    """Procura, nas primeiras linhas, a linha de cabeçalho da TS.

    Retorna (indice_linha, mapa_coluna->indice) ou None.
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        normalizados = {_norm(c): i for i, c in enumerate(row) if c is not None}
        mapa = {}
        for logico, sinonimos in _COLUNAS.items():
            for s in sinonimos:
                if s in normalizados:
                    mapa[logico] = normalizados[s]
                    break
        if all(c in mapa for c in _OBRIGATORIAS):
            return idx + 1, mapa  # openpyxl é 1-based
    return None


def carregar_ts(caminho: str) -> PlanilhaTS:
    """Lê o arquivo .xlsx e devolve a planilha TS reconhecida.

    Varre todas as abas e usa a primeira que contém as colunas obrigatórias.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    for ws in wb.worksheets:
        achado = _achar_cabecalho(ws)
        if not achado:
            continue
        linha_cab, mapa = achado
        linhas = []
        for row in ws.iter_rows(min_row=linha_cab + 1, values_only=True):
            def get(col):
                i = mapa.get(col)
                return row[i] if i is not None and i < len(row) else None

            nome = get("nome")
            mes_key = _mes_key(get("mes"))
            gp = get("gp")
            if nome is None and gp is None and mes_key is None:
                continue  # linha vazia
            if mes_key is None or gp is None:
                continue  # sem mês/GP não serve para o rateio
            id_col = get("id")
            linhas.append(
                {
                    "id": str(id_col).strip() if id_col is not None else "",
                    "nome": str(nome).strip() if nome is not None else "",
                    "mes_key": mes_key,
                    "gp": gp,
                    "horas": _to_float(get("horas")),
                    "proporcao": _to_float(get("proporcao")),
                }
            )
        if linhas:
            return PlanilhaTS(linhas, ws.title)
    raise ValueError(
        "Não encontrei uma aba com as colunas da planilha TS "
        "(Nome Colaborador, Mês, GP, Horas Trabalhadas e Proporção de Hora)."
    )


# ---------------------------------------------------------------------------
# Cálculo do rateio de Plano de Saúde
# ---------------------------------------------------------------------------
def calcular_plano_saude(
    ts: PlanilhaTS,
    mes_key: str,
    segurados: list[dict],
    valor_boleto: float,
) -> dict:
    """Executa o rateio do plano de saúde por GP.

    Parâmetros
    ----------
    ts          : planilha TS carregada.
    mes_key     : 'AAAA-MM' do boleto.
    segurados   : lista de {id, nome, valor} (valor já inclui dependentes).
    valor_boleto: valor total do boleto a ser rateado.

    Retorno: dicionário com a tabela final, tabelas auxiliares e verificações.
    """
    valor_boleto = float(valor_boleto)

    # Índice de segurados por id e por nome (para casar com a TS).
    por_id = {}
    por_nome = {}
    total_segurados = 0.0
    for s in segurados:
        v = _to_float(s.get("valor"))
        total_segurados += v
        if s.get("id"):
            por_id[str(s["id"]).strip()] = v
        if s.get("nome"):
            por_nome[_norm(s["nome"])] = v

    def valor_do_segurado(linha) -> float | None:
        if linha["id"] and linha["id"] in por_id:
            return por_id[linha["id"]]
        n = _norm(linha["nome"])
        if n in por_nome:
            return por_nome[n]
        return None

    # Tabela temporária 2: linhas da TS dos segurados no mês escolhido,
    # já com o valor rateado da linha (= valor do segurado × proporção da hora).
    temp2 = []
    segurados_com_horas = set()
    for l in ts.linhas:
        if l["mes_key"] != mes_key:
            continue
        v = valor_do_segurado(l)
        if v is None:
            continue
        segurados_com_horas.add(l["id"] or _norm(l["nome"]))
        temp2.append(
            {
                "id": l["id"],
                "nome": l["nome"],
                "gp": l["gp"],
                "horas": l["horas"],
                "proporcao": l["proporcao"],
                "valor_segurado": v,
                "valor_linha": v * l["proporcao"],
            }
        )

    # Passo 5: VALOR por GP (SOMASES sobre valor_linha).
    valor_por_gp: dict = {}
    horas_por_gp: dict = {}
    for r in temp2:
        valor_por_gp[r["gp"]] = valor_por_gp.get(r["gp"], 0.0) + r["valor_linha"]
        horas_por_gp[r["gp"]] = horas_por_gp.get(r["gp"], 0.0) + r["horas"]

    total_valor = sum(valor_por_gp.values())

    # Passos 3, 6 e 7: tabela final por GP.
    tabela_final = []
    for gp in sorted(valor_por_gp.keys(), key=lambda g: (str(type(g)), g)):
        valor = valor_por_gp[gp]
        prop = (valor / total_valor) if total_valor else 0.0
        tabela_final.append(
            {
                "gp": gp,
                "horas": round(horas_por_gp.get(gp, 0.0), 4),
                "valor": round(valor, 2),
                "proporcao": prop,
                "valor_final": round(valor_boleto * prop, 2),
            }
        )

    # Ajuste de centavos: garante que a soma do VALOR FINAL feche no boleto.
    soma_final = sum(r["valor_final"] for r in tabela_final)
    dif = round(valor_boleto - soma_final, 2)
    if tabela_final and dif != 0:
        # Joga a diferença de arredondamento no GP de maior valor.
        maior = max(tabela_final, key=lambda r: r["valor_final"])
        maior["valor_final"] = round(maior["valor_final"] + dif, 2)

    # Segurados informados que não tinham horas no mês (não foram rateados).
    sem_horas = [
        s["nome"]
        for s in segurados
        if (str(s.get("id", "")).strip() or _norm(s.get("nome", "")))
        not in segurados_com_horas
    ]

    return {
        "mes_key": mes_key,
        "valor_boleto": round(valor_boleto, 2),
        "total_segurados": round(total_segurados, 2),
        "total_valor_rateado": round(total_valor, 2),
        "diferenca_boleto_segurados": round(valor_boleto - total_segurados, 2),
        "qtd_gps": len(tabela_final),
        "qtd_segurados": len(segurados),
        "segurados_sem_horas": sem_horas,
        "temp2": temp2,
        "tabela_final": tabela_final,
    }


# ---------------------------------------------------------------------------
# Exportação
# ---------------------------------------------------------------------------
def _sanitizar(nome: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(nome)).strip()


def nome_arquivo_saida(mes_key: str, seguradora: str, codigo_boleto: str) -> str:
    """Gera o nome do arquivo no padrão Ano-Mês-Seguradora-codigo_boleto.

    Exemplo: 25-01-Bradesco-BoletoX123.xlsx
    """
    ano, mes = mes_key.split("-")
    return (
        f"{ano[2:]}-{mes}-{_sanitizar(seguradora)}-{_sanitizar(codigo_boleto)}.xlsx"
    )


def exportar_xlsx(resultado: dict, meta: dict, caminho_saida: str) -> str:
    """Gera o arquivo .xlsx com a tabela final e abas de auditoria."""
    wb = openpyxl.Workbook()

    # --- Aba principal: Rateio ---
    ws = wb.active
    ws.title = "Rateio"

    ws.append(["Rateio de Plano de Saúde por GP"])
    ws.append(["Seguradora", meta.get("seguradora", "")])
    ws.append(["Código do boleto", meta.get("codigo_boleto", "")])
    ws.append(["Mês", resultado["mes_key"]])
    ws.append(["Valor do boleto", resultado["valor_boleto"]])
    ws.append(["Soma dos segurados", resultado["total_segurados"]])
    ws.append([])

    cab_linha = ws.max_row + 1
    ws.append(["GP", "HORAS", "VALOR", "PROPORÇÃO", "VALOR FINAL"])
    for r in resultado["tabela_final"]:
        ws.append([r["gp"], r["horas"], r["valor"], r["proporcao"], r["valor_final"]])

    # Linha de total.
    tot_linha = ws.max_row + 1
    ws.append(
        [
            "TOTAL",
            round(sum(r["horas"] for r in resultado["tabela_final"]), 4),
            round(sum(r["valor"] for r in resultado["tabela_final"]), 2),
            round(sum(r["proporcao"] for r in resultado["tabela_final"]), 6),
            round(sum(r["valor_final"] for r in resultado["tabela_final"]), 2),
        ]
    )

    # Formatação numérica.
    for row in ws.iter_rows(min_row=cab_linha + 1, max_row=tot_linha):
        row[2].number_format = "#,##0.00"  # VALOR
        row[3].number_format = "0.0000%"   # PROPORÇÃO
        row[4].number_format = "#,##0.00"  # VALOR FINAL
    ws["B5"].number_format = "#,##0.00"
    ws["B6"].number_format = "#,##0.00"

    # Negrito no título, cabeçalho e total.
    from openpyxl.styles import Font

    ws["A1"].font = Font(bold=True, size=13)
    for c in ws[cab_linha]:
        c.font = Font(bold=True)
    for c in ws[tot_linha]:
        c.font = Font(bold=True)

    larguras = [14, 12, 16, 14, 16]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Aba de auditoria: detalhe por segurado/GP (tabela temporária 2) ---
    ws2 = wb.create_sheet("Detalhe_Segurados")
    ws2.append(
        [
            "Id Colaborador",
            "Nome Colaborador",
            "GP",
            "Horas Trabalhadas",
            "Proporção de Hora",
            "Valor Segurado",
            "Valor Rateado (Valor×Prop.)",
        ]
    )
    for r in resultado["temp2"]:
        ws2.append(
            [
                r["id"],
                r["nome"],
                r["gp"],
                r["horas"],
                r["proporcao"],
                round(r["valor_segurado"], 2),
                round(r["valor_linha"], 2),
            ]
        )
    for c in ws2[1]:
        c.font = Font(bold=True)
    for col, w in zip("ABCDEFG", [16, 28, 10, 16, 16, 14, 22]):
        ws2.column_dimensions[col].width = w

    wb.save(caminho_saida)
    return caminho_saida
